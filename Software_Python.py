import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import time
######################################### Bài làm ####################
rows = int()
columns = int()
cot = int()
hang = int()
So_TT = int()


def Rut_Gon(f1):

    global rows
    global columns
    global hang
    global cot
    global So_TT

    rows = f1.shape[0]              # Gán số dòng vào biến rows
    columns = f1.shape[1]           # Gán só dòng vào biến columns

    # Tìm kiếm hàng và cột nơi có bắt đầu bảng trạng thái
    for i in range(0, rows):
        check = 0
        hang = i
        for y in range(0, columns):
            cot = y
            a = f1.iloc[i, y]                   # Lấy giá trị hàng i, cột y
            if type(a) != str:                  # Nếu cột không phải kiểu chuỗi, bỏ qua
                continue
            a = a.strip()                       # Loại bỏ khoảng trắng hai bên của a
            if (a[0] == 'S' or a[0] == 's') and (ord(a[1]) in range(48, 58)):
                check = 1
                break
        if check == 1:
            break

        # Loại bỏ các dòng và cột dư thừa
    while(cot != 0):
        f1 = f1.drop(f1.columns[0], axis=1)
        cot = cot - 1

    count = 0                           # Số hàng đã xóa

    for i in range(0, hang):
        a = f1.iloc[i-count, 0]
        if type(a) != str:
            f1 = f1.drop([i])
            count = count + 1
            continue
        a = a.strip()
        if a[0] == 'p' or a[0] == 'P' or a[0] == 't' or a[0] == 'T':
            break

    rows = f1.shape[0]              # Cập nhật lại số hàng có trong bài
    columns = f1.shape[1]           # Cập nhật lại số cột có trong bài
    hang = hang - count

    dem = count
    done = 2

    l = list()                      # Nơi chứa các dòng đã xóa

    while (done > 0):
        # Đi vào rút gọn bảng trại thái
        Dict = dict()                       # Từ điển chứa những trạng thái đã duyệt
        k = Dict.items()

        change = dict()                     # Từ điển chứa những trạng thái tương tự
        c = change.items()

        check = 1

        m = 0

        dem = count

        for i in range(hang, rows):
            if i in l:
                m = m + 1
                dem = dem + 1
                continue
            Xoa = 0
            value = ""
            key = ""
            for y in range(0, columns):
                # Lấy giá trị hàng i, cột y
                a = str(f1.iloc[i-m, y])
                a = a.strip()                       # Loại bỏ khoảng trắng hai bên của a
                if y == 0:
                    key = a
                    continue
                value = value + a
                if (y != columns - 1):
                    value = value + ', '
            for item in k:
                if item[1] == value:               # Nếu đã tồn tại giá trị
                    a = str(f1.iloc[i-m, 0])
                    a = a.strip()
                    change[a] = item[0]

                    l.append(i)

                    f1 = f1.drop([i - m + dem])

                    Xoa = 1
            if Xoa == 0:
                Dict[key] = value
            else:
                dem = dem + 1
                m = m + 1
                check = 0

        if (check == 1):
            done = done - 1

        m = 0
        dem = count
        for i in range(hang, rows):

            if i in l:
                m = m + 1
                dem = dem + 1
                continue
            for y in range(0, columns):
                for item in c:
                    a = str(f1.iloc[i-m, y])
                    a = a.strip()
                    if item[0] == a[0:2]:
                        a = a.replace(a[0:2], item[1], 1)
                        f1.iloc[i-m, y] = a

    So_TT = f1.shape[0] - hang

    #f1.drop(f1.filter(regex="Unnamed"), axis=1, inplace=True)
    return f1


def Ma_Hoa_Nhi_Phan(f1, bit):

    TT = list()
    def getbinary(x, n): return format(x, 'b').zfill(n)

    f1['Mã hóa Nhi Phan'] = ''

    for i in range(0, So_TT):
        a = getbinary(i, bit)
        f1.iloc[i + hang, columns] = a

    return f1


def Ma_Hoa_Hot_one(f1):

    f1['Mã hóa hot-one'] = ''

    def getbinary(x, n): return format(x, 'b').zfill(n)

    for i in range(0, So_TT):
        a = 1 << i
        a = getbinary(a, So_TT)
        f1.iloc[i+hang, columns] = a

    return f1


def Ma_Hoa(f1):

    print('\n', 'Bạn muốn mã hóa theo:')
    print('(1): Mã hóa nhị phân')
    print('(2): Mã hóa hot one')
    check = int(input())

    # Tìm kiếm số bits cần thiết

    bit = 1
    while (True):
        a = 2**bit
        if So_TT <= a:
            break
        bit = bit + 1

    if check == 2:
        Ma_Hoa_Hot_one(f1)
    else:
        print('Có số trạng thái là: ', So_TT)
        print('--> Vì vậy số bit cần: ', bit, '\n')
        Ma_Hoa_Nhi_Phan(f1, bit)

    return f1


def LamDep():
    wb = load_workbook('Output.xlsx')

# Chọn sheet cần xóa ô
    b = 0
    ws = wb['Sheet1']
    # Duyệt qua tất cả các ô trong sheet
    for row in ws.iter_rows(1, 2, 1, 6):
        for cell in row:
            for i in range(b, 10):
                a = str(i)
                if cell.value == "Unnamed: " + a:
                    b = i
                    # if cell.value == "Unnamed: 1":
                    # Xóa giá trị của ô
                    cell.value = None
    ws.delete_cols(1, 1)
    if all(cell.value is None for cell in ws[1]):
        # Xóa hàng
        ws.delete_rows(1, 1)
    wb.save('Output.xlsx')


def main():
    print("Nhập địa chỉ file Excel: ")
    DUONG_DAN = str(input())

    print('File: ', DUONG_DAN, '\n')

    xw.Book(DUONG_DAN)
    f1 = pd.read_excel(DUONG_DAN)

    print("Bảng trạng thái ban đầu:", '\n' * 1)
    print(f1, '\n' * 2)

    print("Bảng trạng thái sau khi rút gọn: ", '\n' * 1)
    start = time.time()
    f1 = Rut_Gon(f1)
    LamDep()
    end = time.time()
    print("elapsed_time:{0}".format(end-start))
    print(f1, '\n' * 2)

    print('Bạn có muốn mã hóa trạng thái không ?')
    print('(1): có')
    print('(0): Không')
    check = int(input())

    if check == 1:
        f1 = Ma_Hoa(f1)
        print('Bảng trạng thái sau khi mã hóa: ', '\n', f1, '\n')
    f1.to_excel('Output.xlsx')
    LamDep()
    #f1 = pd.read_excel('Output.xlsx', skiprows=1)
    # f1.to_excel('Output.xlsx')
    print('\n', "-" * 20, 'Kết thúc chương trình', '-' * 20)
    xw.Book('Output.xlsx')


if __name__ == "__main__":
    main()
